import openpyxl
import sys
import csv
import re


def remove_special_characters(text):
    return re.sub(r'[^A-Za-z0-9,]+', '', text)


# Datei einlesen
excel_file = 'src/data/datensaetze.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws = wb.active

rows = []

# Spaltentitel umformatieren so dass sie als SQL Spaltennamen verwendet werden können
first_row = next(ws.iter_rows(values_only=True))
cleaned_first_row = [remove_special_characters(str(cell)) if cell is not None else cell for cell in first_row]
rows.append(("DatasetID", "NodeID") + tuple(cleaned_first_row[56:]))

# Erste beiden Spalten befüllen
col1 = 0
col2 = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is not None:
        col1 = row[0]
        col2 = row[1]
    rows.append((col1, col2) + row[56:])

csv_file = sys.stdout
csv_writer = csv.writer(csv_file)
for row in rows:
    csv_writer.writerow(row)

